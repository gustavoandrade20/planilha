from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
worksheet: Worksheet = workbook.active


sheet_name = 'Minha planilha'

workbook.create_sheet(sheet_name, 0)

worksheet: Worksheet = workbook[sheet_name]

workbook.remove(workbook['Sheet'])

worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    ['João',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

for student in students:
    worksheet.append(student)   

workbook.save(WORKBOOK_PATH)